# -*- coding: utf-8 -*-
"""
core.py — โมดูล logic หลัก
แยกออกจาก GUI เพื่อให้ทดสอบง่ายและแก้ไขได้โดยไม่กระทบหน้าจอ
"""

import re
import shutil
import datetime
from pathlib import Path

from pypdf import PdfReader
import openpyxl
from openpyxl.utils import get_column_letter

# ========== โครงสร้างชีท KBANK (ปรับได้ถ้า format Excel เปลี่ยน) ==========
MERCHANT_ID_ROW = 4      # แถวที่เก็บ Merchant ID ของแต่ละบล็อก
FIRST_DATA_ROW  = 8      # แถวของวันที่ 1 ของเดือน
FIRST_BLOCK_COL = 3      # คอลัมน์แรกของบล็อก (C)
BLOCK_WIDTH     = 4      # ความกว้างต่อ 1 บล็อก (COMM, VAT, NET, SALE)

COL_OFFSET = {"COMM": 0, "VAT": 1, "NET": 2, "SALE": 3}


# ─────────────────────────────────────────────
#  1. อ่าน PDF
# ─────────────────────────────────────────────

def extract_invoice_data(pdf_path: Path, password: str = "") -> dict:
    """อ่าน PDF 1 ใบ คืน dict ข้อมูล หรือ raise ValueError ถ้าผิดพลาด"""
    reader = PdfReader(str(pdf_path))

    if reader.is_encrypted:
        result = reader.decrypt(password)
        if result == 0:
            raise ValueError("รหัสผ่านไม่ถูกต้อง")

    text = "\n".join(page.extract_text() or "" for page in reader.pages)

    # Merchant ID (15 หลัก)
    m = re.search(r"\b(\d{15})\b", text)
    if not m:
        raise ValueError("ไม่พบ Merchant ID")
    merchant_id = m.group(1)

    # วันที่ออกเอกสาร DD/MM/YYYY
    m = re.search(r"(\d{2})/(\d{2})/(\d{4})", text)
    if not m:
        raise ValueError("ไม่พบวันที่ออกเอกสาร")
    day, month, year = m.groups()
    issued_date = datetime.date(int(year), int(month), int(day))

    # แถวสรุปยอด: ITEM  AMOUNT  FEE  VAT  NET
    m = re.search(
        r"(\d+)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})",
        text,
    )
    if not m:
        raise ValueError("ไม่พบแถวสรุปยอดเงิน")

    to_float = lambda s: float(s.replace(",", ""))
    item, amount, fee, vat, net = m.groups()

    return {
        "file": pdf_path.name,
        "merchant_id": merchant_id,
        "issued_date": issued_date,
        "item": int(item),
        "amount": to_float(amount),   # SALE
        "fee": to_float(fee),          # COMM
        "vat": to_float(vat),
        "net": to_float(net),
    }


# ─────────────────────────────────────────────
#  2. หาตำแหน่งใน Excel
# ─────────────────────────────────────────────

def find_merchant_block(ws, merchant_id: str) -> int | None:
    """หา column index ของบล็อก Merchant ID ที่ต้องการ คืน None ถ้าหาไม่เจอ"""
    col = FIRST_BLOCK_COL
    while col <= ws.max_column:
        val = ws.cell(MERCHANT_ID_ROW, col).value
        if val is not None and str(val).strip() == merchant_id:
            return col
        col += BLOCK_WIDTH
    return None


# ─────────────────────────────────────────────
#  3. ประมวลผลหลายไฟล์ (พร้อม callback แจ้ง progress)
# ─────────────────────────────────────────────

def process_pdfs(
    pdf_paths: list[Path],
    excel_path: Path,
    sheet_name: str,
    password: str = "",
    progress_cb=None,   # fn(current, total, message)
) -> dict:
    """
    ประมวลผลทุก PDF แล้วเขียนลง Excel
    คืน dict: {results: [...], success: int, failed: int, backup_path: Path}
    """
    results = []

    # --- backup ---
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = excel_path.parent / "logs" / f"{excel_path.stem}_backup_{ts}.xlsx"
    backup_path.parent.mkdir(exist_ok=True)
    shutil.copy2(excel_path, backup_path)

    wb = openpyxl.load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"ไม่พบชีท '{sheet_name}' ในไฟล์ Excel")
    ws = wb[sheet_name]

    total = len(pdf_paths)
    success = failed = 0

    for i, pdf_path in enumerate(pdf_paths):
        if progress_cb:
            progress_cb(i, total, f"กำลังอ่าน: {pdf_path.name}")

        row = {"file": pdf_path.name, "status": "", "detail": ""}

        try:
            inv = extract_invoice_data(pdf_path, password)

            block_col = find_merchant_block(ws, inv["merchant_id"])
            if block_col is None:
                raise ValueError(f"ไม่พบ Merchant ID {inv['merchant_id']} ในชีท {sheet_name}")

            target_row = FIRST_DATA_ROW + (inv["issued_date"].day - 1)
            ws.cell(target_row, block_col + COL_OFFSET["COMM"]).value = inv["fee"]
            ws.cell(target_row, block_col + COL_OFFSET["VAT"]).value  = inv["vat"]
            ws.cell(target_row, block_col + COL_OFFSET["NET"]).value  = inv["net"]
            ws.cell(target_row, block_col + COL_OFFSET["SALE"]).value = inv["amount"]

            row.update({
                "status": "✅ สำเร็จ",
                "merchant_id": inv["merchant_id"],
                "date": inv["issued_date"].strftime("%d/%m/%Y"),
                "item": inv["item"],
                "amount": inv["amount"],
                "fee": inv["fee"],
                "vat": inv["vat"],
                "net": inv["net"],
                "detail": f"แถว {target_row} คอลัมน์ {get_column_letter(block_col)}",
            })
            success += 1

        except Exception as e:
            row.update({"status": "❌ ล้มเหลว", "detail": str(e)})
            failed += 1

        results.append(row)

    wb.save(excel_path)

    if progress_cb:
        progress_cb(total, total, "เสร็จสิ้น")

    return {
        "results": results,
        "success": success,
        "failed": failed,
        "backup_path": backup_path,
        "timestamp": ts,
    }


# ─────────────────────────────────────────────
#  4. ปลดรหัสผ่าน PDF และบันทึกไฟล์ใหม่
# ─────────────────────────────────────────────

def strip_pdf_password(
    pdf_paths: list,
    password: str,
    output_dir,
    progress_cb=None,
) -> dict:
    """
    ปลดรหัสผ่านออกจาก PDF แล้วบันทึกเป็นไฟล์ใหม่ใน output_dir
    ชื่อไฟล์ใหม่ = ชื่อเดิม + _unlocked.pdf
    คืน dict: {results: [...], success: int, failed: int}
    """
    from pypdf import PdfWriter

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    results = []
    success = failed = 0
    total = len(pdf_paths)

    for i, pdf_path in enumerate(pdf_paths):
        pdf_path = Path(pdf_path)
        if progress_cb:
            progress_cb(i, total, f"กำลังปลดรหัส: {pdf_path.name}")

        row = {"file": pdf_path.name, "status": "", "output": "", "detail": ""}
        try:
            reader = PdfReader(str(pdf_path))

            if reader.is_encrypted:
                result = reader.decrypt(password)
                if result == 0:
                    raise ValueError("รหัสผ่านไม่ถูกต้อง")

            writer = PdfWriter()
            for page in reader.pages:
                writer.add_page(page)
            # ไม่ encrypt — บันทึกแบบไม่มีรหัส

            stem = pdf_path.stem.replace("_unlocked", "")
            out_name = f"{stem}_unlocked.pdf"
            out_path = output_dir / out_name

            with open(out_path, "wb") as f:
                writer.write(f)

            row.update({
                "status": "✅ สำเร็จ",
                "output": out_name,
                "detail": str(out_path),
            })
            success += 1

        except Exception as e:
            row.update({"status": "❌ ล้มเหลว", "detail": str(e)})
            failed += 1

        results.append(row)

    if progress_cb:
        progress_cb(total, total, "เสร็จสิ้น")

    return {"results": results, "success": success, "failed": failed,
            "output_dir": str(output_dir)}
